import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import io

st.set_page_config(
    page_title="📊 გაყიდვების ანალიტიკა",
    page_icon="📊",
    layout="wide"
)

st.markdown("# 📊 გაყიდვების ანალიტიკის სისტემა - XLSX")

with st.sidebar:
    uploaded_file = st.file_uploader(
        "📁 XLSX ფაილის ატვირთვა",
        type=["xlsx", "xls"]
    )

if uploaded_file is not None:
    try:
        # openpyxl-ით ფაილის წაკითხვა
        uploaded_file.seek(0)
        wb = load_workbook(uploaded_file)
        sheet_names = wb.sheetnames
        
        st.success(f"✅ წარმატებით ჩატვირთა! Sheet-ები: {sheet_names}")
        
        # Sheet არჩევა
        selected_sheet = st.selectbox("📊 აირჩიეთ Sheet:", sheet_names)
        
        # pandas-ით მონაცემი
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=0)
        
        st.info(f"📋 რიგი: {len(df)} | 📊 სვეტი: {len(df.columns)}")
        
        # Tabs
        tab1, tab2, tab3 = st.tabs(["📈 დაშბორდი", "📊 ანალიზი", "🔮 პროგნოზა"])
        
        # Tab 1: დაშბორდი
        with tab1:
            st.markdown("## 📊 დაშბორდი")
            
            col1, col2, col3 = st.columns(3)
            col1.metric("📋 რიგი", len(df))
            col2.metric("📊 სვეტი", len(df.columns))
            col3.metric("✅ მონაცემი", df.notna().sum().sum())
            
            st.markdown("### 📋 მონაცემთა ცხრილი")
            st.dataframe(df, use_container_width=True, height=400)
            
            st.markdown("### 📊 სვეტების ინფორმაცია")
            info_df = pd.DataFrame({
                "სვეტი": df.columns,
                "ტიპი": [str(t) for t in df.dtypes],
                "ცარიელი": df.isnull().sum().values
            })
            st.dataframe(info_df, use_container_width=True)
        
        # Tab 2: ანალიზი
        with tab2:
            st.markdown("## 📊 ანალიზი")
            
            selected_cols = st.multiselect(
                "აირჩიეთ სვეტები:",
                df.columns.tolist(),
                default=df.columns.tolist()[:3] if len(df.columns) > 0 else []
            )
            
            if selected_cols:
                st.markdown(f"### აირჩეული სვეტები ({len(selected_cols)})")
                st.dataframe(df[selected_cols], use_container_width=True)
                
                # რიცხვითი ანალიზი
                numeric_cols = df[selected_cols].select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    st.markdown("### 📊 სტატისტიკა")
                    st.dataframe(df[numeric_cols].describe(), use_container_width=True)
            else:
                st.info("📋 სვეტი აირჩიეთ")
        
        # Tab 3: პროგნოზა
        with tab3:
            st.markdown("## 🔮 პროგნოზა")
            
            numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
            
            if numeric_cols:
                col1, col2 = st.columns(2)
                
                with col1:
                    metric = st.selectbox("აირჩიეთ მეტრიკა:", numeric_cols)
                
                with col2:
                    growth = st.slider("ზრდის ტემპი (%):", -100, 200, 10, step=5)
                
                st.markdown(f"### {metric} - პროგნოზა")
                
                current = df[metric].sum()
                forecast = current * (1 + growth / 100)
                diff = forecast - current
                
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("ჯამი", f"{current:,.0f}")
                col2.metric("პროგნოზა", f"{forecast:,.0f}")
                col3.metric("განსხვავება", f"{diff:+,.0f}")
                col4.metric("ზრდა", f"{growth}%")
                
                st.markdown("---")
                
                # დეტალი
                stats = {
                    "მინიმუმი": df[metric].min(),
                    "მაქსიმუმი": df[metric].max(),
                    "საშუო": df[metric].mean(),
                    "მედიანა": df[metric].median()
                }
                
                st.markdown("### 📊 დეტალური სტატისტიკა")
                for key, val in stats.items():
                    st.write(f"**{key}:** {val:,.0f}")
            
            else:
                st.warning("⚠️ რიცხვითი სვეტი არ მოიძებნა")
    
    except Exception as e:
        st.error(f"❌ შეცდომა: {str(e)}")
        st.info(f"💡 დიაგნოსტიკა: {type(e).__name__}")

else:
    st.markdown("## 📁 XLSX ფაილი ატვირთეთ დასაწყებად")
    st.markdown("""
    ### ✅ მხარდაჭერილი ფორმატი:
    - `.xlsx` ფაილი
    - `.xls` ფაილი
    - მრავალი Sheet
    - რიცხვითი, ტექსტური, თარიღის მონაცემი
    """)

st.markdown("---")
st.markdown("© 2025 გაყიდვების ანალიტიკა | **gigalabichi1**")
